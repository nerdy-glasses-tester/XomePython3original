from openpyxl import load_workbook
import utilities.custom_logger as cl
import logging

class ExcelUtils():

    log = cl.customLogger(logging.DEBUG)

    def __init__(self, driver):
        self.driver = driver

    def get_input_rows(self, datafile, testName):

        wb = load_workbook(filename=datafile)

        sheet = wb.active
        self.log.info(sheet)

        sheet = wb.get_sheet_by_name(testName)
        self.log.info("Got active sheet.")
        self.log.info(sheet)
        column_list = []
        row_list = []
        pair = {}

        # Get first row as headers
        rows1 = sheet.iter_rows(min_row=1, max_row=1)
        firstrowheaders = next(rows1)
        for c in firstrowheaders:
            column_list.append(c.value)

        # Get second row as the headers' values
        rows2 = sheet.iter_rows(min_row=2, max_row=2)
        secondrowvalues = next(rows2)
        for d in secondrowvalues:
            row_list.append(d.value)

        # Combine the two list to create a dictionary for key:value lookup
        pair = dict(zip(column_list, row_list))

        return pair
